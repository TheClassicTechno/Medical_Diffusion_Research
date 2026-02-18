#!/usr/bin/env python3
"""
2020 data: Single-delay ASL Pre Diamox only, presurgery from xlsx, subject-level split (no leakage).
- Reads Files 2020.xlsx: subject ID column + CBF_Single_Delay_Pre_Diamox + CBF_Single_Delay_Post_Diamox.
- Keeps only subjects with both Pre and Post = Yes (for bundled pre->post).
- Paths: .../pre_surgery_yes_diamox/perf/asl_single_delay_pre_diamox/CBF_Single_Delay_Pre_Diamox_standard_lin.nii.gz
         .../pre_surgery_yes_diamox/perf/asl_single_delay_post_diamox/ (same CBF filename in that folder; no partial volumes).
- Split by subject ID so no subject appears in more than one of train/val/test.
Requires: pip install openpyxl (for xlsx).
"""
from __future__ import annotations

import os
import random
import json
import argparse
from pathlib import Path

def _load_xlsx_openpyxl(xlsx_path: str, sheet_name: str | None = None) -> list[list]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
    else:
        sh = wb.active
        # Numbers export: first sheet is often "Export Summary" with empty rows; use Pre_surgery_yes_diamox if present
        rows = [list(row) for row in sh.iter_rows(values_only=True)]
        header_ok = rows and any(
            c and ("moyamoya" in str(c).lower() or "Year_Patient" in str(c) or "CBF_Single" in str(c))
            for c in (rows[0] if rows else [])
        )
        if not header_ok and "Pre_surgery_yes_diamox" in wb.sheetnames:
            sh = wb["Pre_surgery_yes_diamox"]
            rows = [list(row) for row in sh.iter_rows(values_only=True)]
        return rows
    return [list(row) for row in sh.iter_rows(values_only=True)]


def _load_xlsx_csv_fallback(csv_path: str) -> list[list]:
    import csv
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f))

# Paths under data1/julih
DATA_2020_ROOT = "/data1/julih/moyamoya_2020_nifti"
XLSX_PATH = "/data1/julih/Files 2020.xlsx"
CBF_PRE_SUBDIR = "derived/pre_surgery_yes_diamox/perf/asl_single_delay_pre_diamox"
CBF_POST_SUBDIR = "derived/pre_surgery_yes_diamox/perf/asl_single_delay_post_diamox"
CBF_FILE = "CBF_Single_Delay_Pre_Diamox_standard_lin.nii.gz"  # pre folder; post folder has same filename (content is post-Diamox)


def load_xlsx_subjects_with_single_delay(xlsx_path: str) -> list[tuple[str, bool, bool]]:
    """
    Parse xlsx (or CSV fallback). Return list of (subject_id, has_pre, has_post) for Single-delay ASL only.
    subject_id = first column (e.g. moyamoya_stanford_2020_001).
    has_pre = CBF_Single_Delay_Pre_Diamox column is Yes (case-insensitive).
    has_post = CBF_Single_Delay_Post_Diamox column is Yes.
    """
    xlsx_path = os.path.abspath(xlsx_path)
    csv_path = xlsx_path.rsplit(".", 1)[0] + ".csv"
    rows = []
    if xlsx_path.lower().endswith(".xlsx"):
        try:
            rows = _load_xlsx_openpyxl(xlsx_path)
        except ImportError:
            if os.path.isfile(csv_path):
                rows = _load_xlsx_csv_fallback(csv_path)
            else:
                raise ImportError("pip install openpyxl to read .xlsx, or export Files 2020.xlsx to Files 2020.csv")
    elif xlsx_path.lower().endswith(".csv") or os.path.isfile(csv_path):
        path = csv_path if os.path.isfile(csv_path) and not xlsx_path.lower().endswith(".csv") else xlsx_path
        rows = _load_xlsx_csv_fallback(path)
    if not rows:
        return []
    # Header may be row 0 or row 1 (Numbers export often has section row then column names)
    header_row_idx = 0
    for idx, row in enumerate(rows[:5]):
        header = [str(c).strip() if c is not None else "" for c in (row or [])]
        if any("cbf_single_delay" in h.lower() or "Year_Patient" in h for h in header):
            header_row_idx = idx
            break
    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    # Find column indices
    id_col = 0  # Year_Patient_Session
    pre_col = None
    post_col = None
    for i, h in enumerate(header):
        h_lower = h.lower()
        if "cbf_single_delay_pre_diamox" in h_lower and "post" not in h_lower:
            pre_col = i
        if "cbf_single_delay_post_diamox" in h_lower:
            post_col = i
    if pre_col is None:
        # Fallback: look for .nii.gz in header
        for i, h in enumerate(header):
            if "CBF_Single_Delay_Pre_Diamox" in str(h) and "Post" not in str(h):
                pre_col = i
            if "CBF_Single_Delay_Post_Diamox" in str(h):
                post_col = i
    if pre_col is None or post_col is None:
        raise ValueError(f"Could not find Pre/Post columns in {xlsx_path}. Header: {header[:25]}")
    out = []
    for row in rows[header_row_idx + 1 :]:
        if not row:
            continue
        sid = str(row[id_col]).strip() if row[id_col] is not None else ""
        if not sid or not sid.startswith("moyamoya_stanford_2020_"):
            continue
        pre_val = (row[pre_col] or "").strip().lower() == "yes" if pre_col is not None else False
        post_val = (row[post_col] or "").strip().lower() == "yes" if post_col is not None else False
        out.append((sid, pre_val, post_val))
    return out


def subject_id_to_folder_name(subject_id: str) -> str:
    """moyamoya_stanford_2020_001 -> moyamoya_stanford_2020_001 (folder name)."""
    return subject_id.strip()


def get_pre_post_paths(data_root: str, subject_id: str) -> tuple[str | None, str | None]:
    """
    Return (pre_path, post_path) for CBF single-delay, no partial volumes.
    Uses standard_lin.nii.gz in pre and post_diamox folders (post folder has same CBF filename).
    """
    folder = subject_id_to_folder_name(subject_id)
    base = os.path.join(data_root, folder)
    pre_dir = os.path.join(base, CBF_PRE_SUBDIR)
    post_dir = os.path.join(base, CBF_POST_SUBDIR)
    pre_path = os.path.join(pre_dir, CBF_FILE)
    # Post folder uses same filename (CBF_Single_Delay_Pre_Diamox_* in post_diamox dir = post-Diamox data)
    post_path = os.path.join(post_dir, CBF_FILE)
    pre_ok = os.path.isfile(pre_path)
    post_ok = os.path.isfile(post_path)
    return (pre_path if pre_ok else None, post_path if post_ok else None)


def build_paired_list(
    data_root: str,
    xlsx_path: str | None = None,
    require_both: bool = True,
    use_xlsx: bool = True,
) -> list[tuple[str, str, str]]:
    """
    (subject_id, pre_path, post_path) for subjects with both Single-delay Pre and Post on disk.
    If use_xlsx and xlsx_path given, filter to subjects where xlsx has Yes for both Pre and Post.
    Otherwise discover from filesystem (all subjects with both paths present).
    """
    if use_xlsx and xlsx_path and os.path.isfile(xlsx_path):
        subjects = load_xlsx_subjects_with_single_delay(xlsx_path)
        candidate_sids = [sid for sid, has_pre, has_post in subjects if has_pre and has_post]
    else:
        candidate_sids = []
        for name in sorted(os.listdir(data_root)):
            if not name.startswith("moyamoya_stanford_2020_") or not os.path.isdir(os.path.join(data_root, name)):
                continue
            candidate_sids.append(name)
    pairs = []
    for sid in candidate_sids:
        pre_path, post_path = get_pre_post_paths(data_root, sid)
        if pre_path and post_path:
            pairs.append((sid, pre_path, post_path))
    return pairs


def subject_level_split(
    pairs: list[tuple[str, str, str]],
    seed: int = 42,
    train_frac: float = 0.75,
    val_frac: float = 0.125,
) -> tuple[list[tuple[str, str, str]], list[tuple[str, str, str]], list[tuple[str, str, str]]]:
    """
    Split by subject ID. Each subject appears in exactly one of train/val/test. No leakage.
    """
    subjects = list({p[0] for p in pairs})
    # Keep stable order per subject; one pair per subject (if multiple scans per subject, all in same split)
    subject_to_pairs = {}
    for p in pairs:
        sid = p[0]
        if sid not in subject_to_pairs:
            subject_to_pairs[sid] = []
        subject_to_pairs[sid].append(p)
    unique_subjects = list(dict.fromkeys(p[0] for p in pairs))  # order preserved, one per subject
    random.seed(seed)
    random.shuffle(unique_subjects)
    n = len(unique_subjects)
    n_train = int(n * train_frac)
    n_val = int(n * val_frac)
    n_test = n - n_train - n_val
    if n_test < 0:
        n_test = 0
        n_val = n - n_train
    train_sids = set(unique_subjects[:n_train])
    val_sids = set(unique_subjects[n_train : n_train + n_val])
    test_sids = set(unique_subjects[n_train + n_val :])
    train_pairs = [p for p in pairs if p[0] in train_sids]
    val_pairs = [p for p in pairs if p[0] in val_sids]
    test_pairs = [p for p in pairs if p[0] in test_sids]
    return train_pairs, val_pairs, test_pairs


def main():
    ap = argparse.ArgumentParser(description="2020 Single-delay Pre/Post from xlsx, subject-level split")
    ap.add_argument("--xlsx", default=XLSX_PATH, help="Path to Files 2020.xlsx")
    ap.add_argument("--data-root", default=DATA_2020_ROOT, help="moyamoya_2020_nifti root")
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--train-frac", type=float, default=0.75)
    ap.add_argument("--val-frac", type=float, default=0.125)
    ap.add_argument("--out", default="/data1/julih/2020_single_delay_split.json", help="Output JSON path")
    ap.add_argument("--list-pairs", action="store_true", help="Print (subject_id, pre_path, post_path) and exit")
    ap.add_argument("--no-xlsx", action="store_true", help="Discover subjects from filesystem only (no xlsx filter)")
    args = ap.parse_args()

    use_xlsx = not args.no_xlsx and os.path.isfile(args.xlsx)
    pairs = build_paired_list(args.data_root, args.xlsx, require_both=True, use_xlsx=use_xlsx)
    print(f"Subjects with both Single-delay Pre and Post (xlsx Yes + on disk): {len(pairs)}")

    if args.list_pairs:
        for sid, pre, post in pairs[:5]:
            print(sid, pre, post)
        if len(pairs) > 5:
            print("...")
        return

    train_pairs, val_pairs, test_pairs = subject_level_split(pairs, seed=args.seed, train_frac=args.train_frac, val_frac=args.val_frac)
    train_sids = sorted({p[0] for p in train_pairs})
    val_sids = sorted({p[0] for p in val_pairs})
    test_sids = sorted({p[0] for p in test_pairs})
    assert len(train_sids) + len(val_sids) + len(test_sids) == len({p[0] for p in pairs}), "Split must partition subjects"
    assert train_sids and (not (set(train_sids) & set(test_sids))), "No leakage"

    out = {
        "seed": args.seed,
        "train_frac": args.train_frac,
        "val_frac": args.val_frac,
        "n_train": len(train_pairs),
        "n_val": len(val_pairs),
        "n_test": len(test_pairs),
        "train_subjects": train_sids,
        "val_subjects": val_sids,
        "test_subjects": test_sids,
        "train": [{"subject_id": p[0], "pre_path": p[1], "post_path": p[2]} for p in train_pairs],
        "val": [{"subject_id": p[0], "pre_path": p[1], "post_path": p[2]} for p in val_pairs],
        "test": [{"subject_id": p[0], "pre_path": p[1], "post_path": p[2]} for p in test_pairs],
    }
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w") as f:
        json.dump(out, f, indent=2)
    print(f"Train {len(train_pairs)} Val {len(val_pairs)} Test {len(test_pairs)}. Written {args.out}")


###############################################################################
# PyTorch Dataset (one input channel, one dataset feed)
###############################################################################

def _load_volume_nifti(path: str, target_size: tuple[int, int, int] | None = None) -> "np.ndarray":
    import nibabel as nib
    import numpy as np
    from scipy.ndimage import zoom as scipy_zoom
    img = nib.load(path)
    vol = np.asarray(img.dataobj).astype(np.float32).squeeze()
    if vol.ndim == 4:
        vol = vol[..., 0]
    if target_size is not None:
        factors = [target_size[i] / vol.shape[i] for i in range(3)]
        vol = scipy_zoom(vol, factors, order=1)
    vmin, vmax = vol.min(), vol.max()
    if (vmax - vmin) > 1e-8:
        vol = (vol - vmin) / (vmax - vmin)
    else:
        vol = np.zeros_like(vol)
    return vol.astype(np.float32)


class Dataset2020SingleDelay:
    """
    One channel (pre CBF), one target (post CBF). Loads from split JSON.
    Optional target_size for resize; no partial volumes (uses CBF only).
    """
    def __init__(
        self,
        split_json: str,
        split_name: str = "train",
        target_size: tuple[int, int, int] | None = (128, 128, 64),
    ):
        with open(split_json) as f:
            data = json.load(f)
        self.items = data[split_name]  # list of {"subject_id", "pre_path", "post_path"}
        self.target_size = target_size

    def __len__(self) -> int:
        return len(self.items)

    def __getitem__(self, idx: int) -> tuple["torch.Tensor", "torch.Tensor"]:
        import torch
        item = self.items[idx]
        pre = _load_volume_nifti(item["pre_path"], self.target_size)
        post = _load_volume_nifti(item["post_path"], self.target_size)
        pre_t = torch.from_numpy(pre).unsqueeze(0).float()   # (1, H, W, D)
        post_t = torch.from_numpy(post).unsqueeze(0).float()
        return pre_t, post_t


if __name__ == "__main__":
    main()
